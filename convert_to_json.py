import os
import glob
import json
import openpyxl

# ====================================================
# 1. 중학교 엑셀 -> JSON 변환 (data/evaluation standard/middle/ 에 바로 저장)
# ====================================================
def convert_middle_to_json():
    middle_dir = os.path.join("data", "evaluation standard", "middle")
    os.makedirs(middle_dir, exist_ok=True)

    target_files = (
        glob.glob(os.path.join(middle_dir, "*.xlsx")) +
        glob.glob(os.path.join(middle_dir, "*.xls"))
    )

    if not target_files:
        print(f"⚠️ [{middle_dir}] 경로에서 엑셀 파일을 찾지 못했습니다.")
        return False

    file_path = target_files[0]
    print(f"📂 [중학교] 엑셀 변환 시작: {file_path}")

    subject_dict = {}

    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
        sheet = wb.active

        for row in sheet.iter_rows(values_only=True):
            if not row or len(row) < 3:
                continue

            col_a, col_b, col_c = row[0], row[1], row[2]

            # 헤더 행 건너뛰기
            if str(col_a).strip() in ['교과', '교과 이름', '교과명', '과목명', 'None']:
                continue

            if not col_a:
                continue

            subject_name = str(col_a).strip()
            code = str(col_b).strip() if col_b is not None else ""
            content = str(col_c).strip() if col_c is not None else ""

            if not content or content == "None":
                continue

            if code and code != "None":
                if not code.startswith("["):
                    code = f"[{code}]"
                entry = f"{code} {content}" if not content.startswith("[") else content
            else:
                entry = content

            if subject_name not in subject_dict:
                subject_dict[subject_name] = []

            subject_dict[subject_name].append(entry)

        # data/evaluation standard/middle/ 폴더에 [과목명.json] 바로 저장
        for subj_name, std_list in subject_dict.items():
            safe_subj_name = subj_name.replace("/", "_").replace("\\", "_")
            out_json_path = os.path.join(middle_dir, f"{safe_subj_name}.json")

            with open(out_json_path, 'w', encoding='utf-8') as f:
                json.dump(std_list, f, ensure_ascii=False, indent=2)

            print(f"  └─ 📄 middle/{safe_subj_name}.json ({len(std_list)}개 성취기준)")

        print("✨ 중학교 성취기준 JSON 변환 완료!\n")
        return True

    except Exception as e:
        print(f"❌ 중학교 엑셀 변환 오류: {e}")
        return False


# ====================================================
# 2. 고등학교 엑셀 -> JSON 변환 (data/evaluation standard/high/[교과]/[세부교과].json)
# ====================================================
def convert_high_to_json():
    high_dir = os.path.join("data", "evaluation standard", "high")
    os.makedirs(high_dir, exist_ok=True)

    target_files = (
        glob.glob(os.path.join(high_dir, "*.xlsx")) +
        glob.glob(os.path.join(high_dir, "*.xls"))
    )

    if not target_files:
        print(f"⚠️ [{high_dir}] 경로에서 엑셀 파일을 찾지 못했습니다.")
        return False

    file_path = target_files[0]
    print(f"📂 [고등학교] 엑셀 변환 시작: {file_path}")

    # high_dict = { "국어": { "공통국어1": [...], "공통국어2": [...] } }
    high_dict = {}

    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
        sheet = wb.active

        for row in sheet.iter_rows(values_only=True):
            if not row or len(row) < 4:
                continue

            col_a, col_b, col_c, col_d = row[0], row[1], row[2], row[3]

            # 헤더 행 건너뛰기
            if str(col_a).strip() in ['교과', '교과 이름', '교과명'] or str(col_b).strip() in ['세부교과', '세부교과 이름', '세부교과명']:
                continue

            if not col_a or not col_b:
                continue

            main_subject = str(col_a).strip()      # A열: 교과명
            detail_subject = str(col_b).strip()    # B열: 세부교과명
            code = str(col_c).strip() if col_c is not None else ""
            content = str(col_d).strip() if col_d is not None else ""

            if not content or content == "None":
                continue

            if code and code != "None":
                if not code.startswith("["):
                    code = f"[{code}]"
                entry = f"{code} {content}" if not content.startswith("[") else content
            else:
                entry = content

            if main_subject not in high_dict:
                high_dict[main_subject] = {}
            if detail_subject not in high_dict[main_subject]:
                high_dict[main_subject][detail_subject] = []

            high_dict[main_subject][detail_subject].append(entry)

        # A열(교과) 폴더 생성 후, 그 안에 B열(세부교과).json 파일 생성
        for main_subj, detail_map in high_dict.items():
            safe_main = main_subj.replace("/", "_").replace("\\", "_")
            main_folder_path = os.path.join(high_dir, safe_main)
            os.makedirs(main_folder_path, exist_ok=True) # A열 교과명 폴더 생성

            for detail_subj, std_list in detail_map.items():
                safe_detail = detail_subj.replace("/", "_").replace("\\", "_")
                out_json_path = os.path.join(main_folder_path, f"{safe_detail}.json")

                with open(out_json_path, 'w', encoding='utf-8') as f:
                    json.dump(std_list, f, ensure_ascii=False, indent=2)

                print(f"  └─ 📁 high/{safe_main}/📄 {safe_detail}.json ({len(std_list)}개 성취기준)")

        print("✨ 고등학교 성취기준 JSON 변환 완료!\n")
        return True

    except Exception as e:
        print(f"❌ 고등학교 엑셀 변환 오류: {e}")
        return False


def convert_all_to_json():
    print("🔄 [전체] 성취기준 JSON 변환 프로세스 시작...\n")
    convert_middle_to_json()
    convert_high_to_json()
    print("🎉 모든 성취기준 JSON 변환 작업이 완료되었습니다!")


if __name__ == '__main__':
    convert_all_to_json()